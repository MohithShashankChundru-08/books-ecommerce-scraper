import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def get_bookdetails(pages):
    scraped_data = []
    for page in range(1,1+pages):
        URL = f'https://books.toscrape.com/catalogue/page-{page}.html'
        try:
            response = requests.get(URL)
            response.raise_for_status()
        except Exception as e:
            print('Something went wrong!:',e)
            return[]
        soup = BeautifulSoup(response.text,'html.parser')
        for tag in soup.find_all('article',class_ = 'product_pod'):
            name_data = tag.find('h3').find('a')
            name = name_data.get('title')
            rating_data = tag.find('p', class_='star-rating').get('class')
            if rating_data[1] == 'One':
                rating = 1
            elif rating_data[1] =='Two':
                rating = 2
            elif rating_data[1] =='Three':
                rating = 3
            elif rating_data[1] =='Four':
                rating = 4
            else:
                rating = 5
            price_data = tag.find('p',class_ = 'price_color').text.strip()
            price = price_data.replace('£','').replace('Â','')
            availability = tag.find('div',class_ = 'product_price').find('p',class_ = 'instock availability').text.strip()
            scraped_data.append({'Book name':name,'Price(£)':float(price),'Rating':rating,'Availability':availability})
    return scraped_data
    
def clean_data(scraped_data):
    df = pd.DataFrame(scraped_data)
    df = df.drop_duplicates()
    df = df.dropna(subset=['Book name'])
    return df

def save_data(df,filename):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'Books Data'
    ws1.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws1.append(list(row))
    header_fill = PatternFill(start_color='4F4F4F',end_color='4F4F4F',fill_type='solid')
    alt_fill = PatternFill(start_color='EFEFEF',end_color='EFEFEF',fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    normal_font = Font(color='000000')
    alignment = Alignment(horizontal='left', vertical='center')
    border_side = Side(style='thin',color='000000')
    border = Border(left=border_side,right=border_side,bottom=border_side,top=border_side)
    for col in range(1,len(df.columns)+1):
        cell = ws1.cell(row=1,column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment
        cell.border = border

    for col in range(1,len(df.columns)+1):
        for row in range(2,len(df)+2):
            cell = ws1.cell(row=row, column=col)
            if row % 2 == 0:
                cell.fill = alt_fill
            cell.font = normal_font
            cell.alignment = alignment
            cell.border = border

    ws1.freeze_panes = 'A2'

    ws2 = wb.create_sheet(title='Summary')
    total_books = len(df['Book name'])
    avg_price = df['Price(£)'].mean()
    rating5 = 0
    rating4 =0
    rating3 =0
    rating2 =0
    rating1 =0
    for rating_details in df['Rating']:
        if rating_details == 5:
            rating5 += 1
        elif rating_details == 4:
            rating4 += 1
        elif rating_details == 3:
            rating3 += 1
        elif rating_details == 2:
            rating2 += 1
        else:
            rating1 += 1

    ws2['A1'] = 'Metric'
    ws2['B1'] = 'Value'
    ws2['A2'] = 'Total Books'
    ws2['B2'] = total_books
    ws2['A3'] = 'Average Price Of The Book'
    ws2['B3'] = avg_price
    ws2['A4'] = 'Number Of Books Having 5 Star Rating'
    ws2['B4'] = rating5
    ws2['A5'] = 'Number Of Books Having 4 Star Rating'
    ws2['B5'] = rating4
    ws2['A6'] = 'Number Of Books Having 3 Star Rating'
    ws2['B6'] = rating3
    ws2['A7'] = 'Number Of Books Having 2 Star Rating'
    ws2['B7'] = rating2
    ws2['A8'] = 'Number Of Books Having 1 Star Rating'
    ws2['B8'] = rating1

    ws3 = wb.create_sheet(title='Top Rated Books')
    sorted_data = df[df['Rating']>3]
    ws3.append(list(sorted_data.columns))
    for row in sorted_data.itertuples(index=False):
        ws3.append(list(row))
    for col in range(1,len(sorted_data.columns)+1):
        cell = ws3.cell(row=1,column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment
        cell.border = border

    for col in range(1,len(sorted_data.columns)+1):
        for row in range(2,len(sorted_data)+2):
            cell = ws3.cell(row=row, column=col)
            if row % 2 == 0:
                cell.fill = alt_fill
            cell.font = normal_font
            cell.alignment = alignment
            cell.border = border
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter 
            for cell in col:
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except:
                    pass
            adjusted_width = max_length 
            ws.column_dimensions[column_letter].width = adjusted_width
    ws3.freeze_panes = 'A2'
    wb.save(filename)
    print("Excel file created and saved successfully.")

def main():
    filename = input('Enter your filename (eg. data.xlsx): ')
    books = get_bookdetails(50)
    cleaned = clean_data(books)
    save_data(cleaned,filename)

if __name__=='__main__':
    main()